#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county.

#TODO Reads the data from the Excel spreadsheet
import openpyxl, pprint
wb = openpyxl.load_workbook('data/censuspopdata.xlsx')
sheet = wb['Population by Census Tract']
print(sheet.cell(row=1, column=2).value)
countyData = {}
#TODO Counts the number of census tracts in each county
for row in range (2, sheet.max_row+1):
    state  = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop    = sheet['D' + str(row)].value 
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done.')

#TODO Prints the results