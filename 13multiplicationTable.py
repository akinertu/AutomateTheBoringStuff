#! python3
# multiplicationTable.py takes a number N and creates an N×N multiplication table in an Excel spreadsheet.

import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']

print('Enter N to create NxN multiplication table.')
N = int(input())

for rowNum in range (1,N+2):
    for colNum in range (1,N+2):
        
        if rowNum < 2 and colNum > 1:
            sheet.cell(row=rowNum, column=colNum).value = colNum-1
            fontObj1 = Font(bold=True)
            sheet.cell(row=rowNum, column=colNum).font = fontObj1
            
        elif rowNum > 1 and colNum < 2:
            sheet.cell(row=rowNum, column=colNum).value = rowNum-1
            sheet.cell(row=rowNum, column=colNum).font = fontObj1
        else:
            sheet.cell(row=rowNum, column=colNum).value = (colNum-1)*(rowNum-1)
sheet.cell(row=1, column=1).value = 'N x N'
  
wb.save('data/multiplicationTable.xlsx')
