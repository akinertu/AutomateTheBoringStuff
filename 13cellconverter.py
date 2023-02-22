#! python3
# cellinverter.py # inverts the row and column of the cells in the spreadsheet.

import openpyxl
wb = openpyxl.load_workbook('data/example2.xlsx')
sheet = wb['Sheet1']
newsheet = wb.create_sheet(index=1, title='inversecopy')

for rowNum in range (1,sheet.max_row+1):
    for colNum in range (1,sheet.max_column+1):
        newsheet.cell(row=colNum, column=rowNum).value = sheet.cell(row=rowNum, column=colNum).value
        
wb.save('data/example2.xlsx')
        